import bs4
import requests
import openpyxl
#pip install bs4 requests openpyxl

#inbuild Library
import re
import sys
import os


url = input("Course URL:- ")
page = requests.get(url)
soup = bs4.BeautifulSoup(page.content, 'html.parser')

try:
    title = soup.find(attrs={"name": "description2"})['value']
    title=re.sub('\W+',' ',title)
    print(title)
except KeyError:
    print("Error 404: Course doesn't exist")
    sys.exit(1)

    
wb = openpyxl.Workbook()
ws1 = wb.active

ws1['A1'].value="Title"
ws1['B1'].value="Topic"
ws1['C1'].value="Description"
ws1['D1'].value="What You Will Learn"


#Get the Description

ws1['A2'].value=title
topics = soup.find(attrs={"name": "course_name"})['value']
ws1['B2'].value=topics



descp = soup.find(attrs={"class": "about-course course-section"}).div

ws1['C2'].value=str(descp)

styles = soup.find_all("style")[3]


bgimg = re.findall(r"background-image:[^\>]*'\);", str(styles))
bgimg = str(re.findall(r'(?:http\:|https\:)?\/\/.*\.(?:png|jpg|webp|jpeg)', str(bgimg))[0])


scripts = soup.find_all("script")[7]
try:
    emvid = re.findall(r"(?:https?:\/\/)?(?:www\.)?youtu\.?be(?:\.com)?\/?.*(?:watch|embed)?(?:.*v=|v\/|\/)([\w\-_]+)\&?",str(scripts))[0]
    video_link = "https://www.youtube.com/embed/"+emvid
except IndexError:
    video_link = "Embeded link not avilable"


cls_dtl = soup.find_all(attrs={"class": "little-h4"})

delivery_method = cls_dtl[1].text
if re.search(r"Online|online",delivery_method):
    delivery_method = "Online"


instruction_type = cls_dtl[0].text

price = soup.find("span", class_= "price-set")
#need javascript fetching module to work on that


content = soup.find_all(attrs={"id": "course-curriculum"})[1]
cnt = re.sub('</? *div[^>]*>|</? *a[^>]*>|Preview|<style([\\s\\S]+?)</style>', '', str(content))
cnt = re.sub(r'\n\s*\n','\n',cnt,re.MULTILINE)


wwl = soup.find_all(attrs={"class": "accordion-content"})[0].get_text()
ws1['D2'].value=wwl

try:
    preq = soup.find('h3', text = "Prerequisites").findNext('ul').get_text()
    ws1['E1'].value="Prerequisites"
except AttributeError:
    ws1.merge_cells('D1:E1')
    preq = None
    
ws1['E2'].value=preq


review = soup.find(attrs={"id": "reviews"})

try:
    name_info = review.find_all(attrs={"class": ["name"]})
    del name_info[1::2]
    comment= review.find_all("p")
    pic = review.find_all("div", class_= "profile-pic")
    for i in range(len(name_info)):
        ws1[chr(ord('F')+i)+str(1)].value = "Review "+ str(i+1)
        ws1[chr(ord('F')+i)+str(2)].value = str(name_info[i].get_text()) + str("\n") + str(comment[i].get_text()) + str("\n")+ str(pic[i].img["data-echo"])
except AttributeError:
    i=0

ws1[chr(ord('F')+i)+str(1)].value = "Cover Photo"
ws1[chr(ord('F')+i)+str(2)].value = bgimg
ws1[chr(ord('F')+i+1)+str(1)].value = "Video Link"
ws1[chr(ord('F')+i+1)+str(2)].value = video_link
ws1[chr(ord('F')+i+2)+str(1)].value = "Delivery Method"
ws1[chr(ord('F')+i+2)+str(2)].value = delivery_method
ws1[chr(ord('F')+i+3)+str(1)].value = "Instruction Type"
ws1[chr(ord('F')+i+3)+str(2)].value = instruction_type
ws1[chr(ord('F')+i+4)+str(1)].value = "Cotent"
ws1[chr(ord('F')+i+4)+str(2)].value = str(cnt)


styled_font = openpyxl.styles.Font(color='00FF0000', italic=True, bold=True)
styled_border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
for cell in ws1["1:1"]:
    cell.font = styled_font
    cell.border = styled_border

ws1.merge_cells('B3:H4')
ws1['B3'].value=title+" \n "+url
ws1['B3'].fill = openpyxl.styles.PatternFill(bgColor="00EF0000", fill_type = "solid")

file_name = title+'.xlsx'
wb.save(filename= file_name)
os.startfile(file_name)
